"""
Professional Excel report generation using openpyxl
"""

from typing import Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelReportGenerator:
    """Generate professional Excel reports for structural analysis"""

    def __init__(self):
        self.wb = Workbook()
        self.title_font = Font(size=16, bold=True, color="FFFFFF")
        self.header_font = Font(size=12, bold=True, color="FFFFFF")
        self.data_font = Font(size=10)
        self.title_fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def create_comprehensive_report(
        self,
        gate_config: Dict,
        analysis_results: Dict,
        material_data: Dict,
        output_path: str,
    ) -> None:
        """Create comprehensive Excel report with multiple worksheets"""

        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Create worksheets
        self._create_summary_sheet(gate_config, analysis_results)
        self._create_calculations_sheet(analysis_results)
        self._create_material_sheet(material_data)
        self._create_loading_sheet(gate_config)
        self._create_design_criteria_sheet()

        # Save workbook
        self.wb.save(output_path)

    def _create_summary_sheet(self, gate_config: Dict, analysis_results: Dict) -> None:
        """Create executive summary worksheet"""
        ws = self.wb.create_sheet("Executive Summary")

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "CANTILEVER SLIDE GATE - STRUCTURAL ANALYSIS REPORT"
        ws["A1"].font = self.title_font
        ws["A1"].fill = self.title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Project information
        ws["A3"] = "PROJECT INFORMATION"
        ws["A3"].font = self.header_font
        ws["A3"].fill = self.header_fill
        ws["A3"].alignment = Alignment(horizontal="center")

        project_data = [
            ["Project Name:", gate_config.get("project_name", "Slide Gate Design")],
            ["Gate Width:", f"{gate_config.get('width_mm', 0) / 1000:.1f} m"],
            ["Gate Height:", f"{gate_config.get('height_mm', 2400) / 1000:.1f} m"],
            ["Design Wind Speed:", f"{gate_config.get('wind_speed_ms', 35)} m/s"],
            ["Analysis Date:", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Engineer:", "Structural Analysis Software"],
            ["Material Grade:", gate_config.get("material_grade", "A572 Grade 50")],
            ["Safety Factor:", f"{gate_config.get('safety_factor', 2.5):.1f}"],
        ]

        for i, (label, value) in enumerate(project_data, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = Font(bold=True)

        # Analysis results summary
        ws["A13"] = "ANALYSIS RESULTS SUMMARY"
        ws["A13"].font = self.header_font
        ws["A13"].fill = self.header_fill
        ws["A13"].alignment = Alignment(horizontal="center")

        results_data = [
            [
                "Maximum Bending Moment:",
                f"{analysis_results.get('max_moment_Nmm', 0) / 1e6:.1f} kN⋅m",
            ],
            [
                "Maximum Bending Stress:",
                f"{analysis_results.get('max_stress_Pa', 0) / 1e6:.1f} MPa",
            ],
            [
                "Allowable Stress:",
                f"{analysis_results.get('allowable_stress_Pa', 0) / 1e6:.1f} MPa",
            ],
            [
                "Stress Utilization Ratio:",
                f"{analysis_results.get('stress_ratio', 0):.2f}",
            ],
            [
                "Maximum Deflection:",
                f"{analysis_results.get('max_deflection_mm', 0):.1f} mm",
            ],
            [
                "Deflection Limit (L/240):",
                f"{analysis_results.get('deflection_limit_mm', 0):.1f} mm",
            ],
            ["Deflection Ratio:", f"{analysis_results.get('deflection_ratio', 0):.2f}"],
            [
                "Design Status:",
                "ADEQUATE"
                if analysis_results.get("safety_adequate", False)
                else "INADEQUATE",
            ],
        ]

        for i, (label, value) in enumerate(results_data, start=14):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = Font(bold=True)

            # Color code the design status
            if "Design Status" in label:
                if "ADEQUATE" in str(value):
                    ws[f"B{i}"].fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid"
                    )
                else:
                    ws[f"B{i}"].fill = PatternFill(
                        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                    )

        # Apply borders and formatting
        for row in ws["A1:B21"]:
            for cell in row:
                cell.border = self.border

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_calculations_sheet(self, analysis_results: Dict) -> None:
        """Create detailed calculations worksheet"""
        ws = self.wb.create_sheet("Detailed Calculations")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "DETAILED STRUCTURAL CALCULATIONS"
        ws["A1"].font = self.title_font
        ws["A1"].fill = self.title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Create data table
        if "positions_mm" in analysis_results:
            positions = analysis_results["positions_mm"]
            moments = analysis_results["moments_Nmm"]
            shears = analysis_results["shears_N"]
            deflections = analysis_results["deflections_mm"]

            # Headers
            headers = [
                "Position (mm)",
                "Moment (kN⋅m)",
                "Shear (kN)",
                "Deflection (mm)",
                "Stress (MPa)",
                "Curvature (1/m)",
            ]
            for i, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=i, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows (sample every 10th point to avoid too much data)
            step = max(1, len(positions) // 100)  # Max 100 data points
            for i in range(0, len(positions), step):
                row = i // step + 4
                ws.cell(row=row, column=1, value=float(positions[i]))
                ws.cell(row=row, column=2, value=float(moments[i] / 1e6))
                ws.cell(row=row, column=3, value=float(shears[i] / 1000))
                ws.cell(row=row, column=4, value=float(deflections[i]))
                # Calculate stress and curvature
                if "section_modulus_mm3" in analysis_results:
                    stress = moments[i] / analysis_results["section_modulus_mm3"]
                    ws.cell(row=row, column=5, value=float(stress))
                if "EI_Nm2" in analysis_results:
                    curvature = (
                        moments[i] / analysis_results["EI_Nm2"] * 1000
                    )  # Convert to 1/m
                    ws.cell(row=row, column=6, value=float(curvature))

        # Apply formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border
                if cell.row > 3:  # Data rows
                    cell.alignment = Alignment(horizontal="right")
                    if cell.column > 1:  # Numeric columns
                        cell.number_format = "0.00"

    def _create_material_sheet(self, material_data: Dict) -> None:
        """Create material properties worksheet"""
        ws = self.wb.create_sheet("Material Properties")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "MATERIAL PROPERTIES AND SPECIFICATIONS"
        ws["A1"].font = self.title_font
        ws["A1"].fill = self.title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Material properties table
        properties = [
            ["Property", "Value", "Unit", "Reference"],
            [
                "Material Grade",
                material_data.get("grade", "A572 Grade 50"),
                "-",
                "ASTM A572",
            ],
            [
                "Yield Strength",
                f"{material_data.get('yield_strength_Pa', 0) / 1e6:.0f}",
                "MPa",
                "ASTM A572",
            ],
            [
                "Ultimate Strength",
                f"{material_data.get('ultimate_strength_Pa', 0) / 1e6:.0f}",
                "MPa",
                "ASTM A572",
            ],
            [
                "Elastic Modulus",
                f"{material_data.get('elastic_modulus_Pa', 0) / 1e9:.0f}",
                "GPa",
                "ASTM A572",
            ],
            [
                "Density",
                f"{material_data.get('density_kg_m3', 0):.0f}",
                "kg/m³",
                "ASTM A572",
            ],
            [
                "Poisson's Ratio",
                f"{material_data.get('poisson_ratio', 0.3):.2f}",
                "-",
                "Typical Value",
            ],
            [
                "Thermal Expansion",
                f"{material_data.get('thermal_expansion', 12e-6) * 1e6:.1f}",
                "×10⁻⁶/°C",
                "Typical Value",
            ],
        ]

        for i, row_data in enumerate(properties, start=3):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center")
                cell.border = self.border

        # Section properties
        ws["A12"] = "SECTION PROPERTIES"
        ws["A12"].font = self.header_font
        ws["A12"].fill = self.header_fill

        if "section" in material_data:
            section = material_data["section"]
            section_props = [
                ["Section Designation", section.get("name", "HSS150x150x6"), "-"],
                ["Cross-sectional Area", f"{section.get('area_mm2', 0):.0f}", "mm²"],
                ["Moment of Inertia (Ix)", f"{section.get('Ix_mm4', 0):.0f}", "mm⁴"],
                ["Moment of Inertia (Iy)", f"{section.get('Iy_mm4', 0):.0f}", "mm⁴"],
                ["Section Modulus (Sx)", f"{section.get('Sx_mm3', 0):.0f}", "mm³"],
                ["Section Modulus (Sy)", f"{section.get('Sy_mm3', 0):.0f}", "mm³"],
                ["Radius of Gyration (rx)", f"{section.get('rx_mm', 0):.1f}", "mm"],
                ["Radius of Gyration (ry)", f"{section.get('ry_mm', 0):.1f}", "mm"],
            ]

            for i, (prop, value, unit) in enumerate(section_props, start=13):
                ws[f"A{i}"] = prop
                ws[f"B{i}"] = value
                ws[f"C{i}"] = unit
                ws[f"A{i}"].font = Font(bold=True)
                for col in ["A", "B", "C"]:
                    ws[f"{col}{i}"].border = self.border

    def _create_loading_sheet(self, gate_config: Dict) -> None:
        """Create loading conditions worksheet"""
        ws = self.wb.create_sheet("Loading Conditions")

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "LOADING CONDITIONS AND LOAD COMBINATIONS"
        ws["A1"].font = self.title_font
        ws["A1"].fill = self.title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Wind loading
        ws["A3"] = "WIND LOADING"
        ws["A3"].font = self.header_font
        ws["A3"].fill = self.header_fill

        wind_data = [
            ["Design Wind Speed", f"{gate_config.get('wind_speed_ms', 35)}", "m/s"],
            [
                "Wind Pressure (q)",
                f"{0.5 * 1.225 * gate_config.get('wind_speed_ms', 35) ** 2:.1f}",
                "Pa",
            ],
            ["Wind Pressure Coefficient", "1.2", "-"],
            [
                "Effective Wind Pressure",
                f"{0.5 * 1.225 * gate_config.get('wind_speed_ms', 35) ** 2 * 1.2:.1f}",
                "Pa",
            ],
            [
                "Gate Area",
                f"{gate_config.get('width_mm', 6000) * gate_config.get('height_mm', 2400) / 1e6:.1f}",
                "m²",
            ],
            [
                "Total Wind Force",
                f"{0.5 * 1.225 * gate_config.get('wind_speed_ms', 35) ** 2 * 1.2 * gate_config.get('width_mm', 6000) * gate_config.get('height_mm', 2400) / 1e6 / 1000:.1f}",
                "kN",
            ],
        ]

        for i, (item, value, unit) in enumerate(wind_data, start=4):
            ws[f"A{i}"] = item
            ws[f"B{i}"] = value
            ws[f"C{i}"] = unit
            ws[f"A{i}"].font = Font(bold=True)
            for col in ["A", "B", "C"]:
                ws[f"{col}{i}"].border = self.border

        # Dead loads
        ws["A11"] = "DEAD LOADS"
        ws["A11"].font = self.header_font
        ws["A11"].fill = self.header_fill

        dead_loads = [
            ["Gate Self Weight", f"{gate_config.get('gate_weight_kg', 500):.0f}", "kg"],
            [
                "Gate Self Weight Force",
                f"{gate_config.get('gate_weight_kg', 500) * 9.81 / 1000:.1f}",
                "kN",
            ],
            [
                "Distributed Dead Load",
                f"{gate_config.get('gate_weight_kg', 500) * 9.81 / gate_config.get('width_mm', 6000):.2f}",
                "N/mm",
            ],
        ]

        for i, (item, value, unit) in enumerate(dead_loads, start=12):
            ws[f"A{i}"] = item
            ws[f"B{i}"] = value
            ws[f"C{i}"] = unit
            ws[f"A{i}"].font = Font(bold=True)
            for col in ["A", "B", "C"]:
                ws[f"{col}{i}"].border = self.border

    def _create_design_criteria_sheet(self) -> None:
        """Create design criteria and code references worksheet"""
        ws = self.wb.create_sheet("Design Criteria")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "DESIGN CRITERIA AND CODE REFERENCES"
        ws["A1"].font = self.title_font
        ws["A1"].fill = self.title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Design codes
        ws["A3"] = "APPLICABLE CODES AND STANDARDS"
        ws["A3"].font = self.header_font
        ws["A3"].fill = self.header_fill

        codes = [
            ["AISC 360", "Specification for Structural Steel Buildings"],
            ["ASCE 7", "Minimum Design Loads for Buildings and Other Structures"],
            ["AWS D1.1", "Structural Welding Code - Steel"],
            ["ASTM A572", "Standard Specification for High-Strength Low-Alloy Steel"],
            [
                "ACI 318",
                "Building Code Requirements for Structural Concrete (foundations)",
            ],
        ]

        for i, (code, description) in enumerate(codes, start=4):
            ws[f"A{i}"] = code
            ws[f"B{i}"] = description
            ws[f"A{i}"].font = Font(bold=True)
            for col in ["A", "B"]:
                ws[f"{col}{i}"].border = self.border

        # Design criteria
        ws["A10"] = "DESIGN CRITERIA"
        ws["A10"].font = self.header_font
        ws["A10"].fill = self.header_fill

        criteria = [
            ["Safety Factor", "2.5", "Applied to yield strength"],
            ["Deflection Limit", "L/240", "Maximum allowable deflection"],
            ["Wind Load Factor", "1.2", "Applied to wind pressure"],
            ["Load Combinations", "1.2D + 1.6W", "ASCE 7 load combinations"],
            ["Connection Safety", "3.0", "Applied to connection design"],
            ["Fatigue Considerations", "Infinite Life", "Gate operation cycles"],
        ]

        for i, (criterion, value, note) in enumerate(criteria, start=11):
            ws[f"A{i}"] = criterion
            ws[f"B{i}"] = value
            ws[f"C{i}"] = note
            ws[f"A{i}"].font = Font(bold=True)
            for col in ["A", "B", "C"]:
                ws[f"{col}{i}"].border = self.border
